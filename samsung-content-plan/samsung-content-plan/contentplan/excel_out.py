"""Запись исправленного контент-плана в Excel.

Дедлайны пишутся ФОРМУЛАМИ, а не значениями: менеджер меняет дату старта —
дедлайны пересчитываются прямо в файле, без запуска пайплайна.

Формула переноса выходных: WORKDAY(старт − N + 1; −1) возвращает последний
рабочий день строго раньше (старт − N + 1), то есть сам день, если он рабочий,
и предыдущую пятницу, если он выпал на выходные.
"""
from __future__ import annotations

import shutil
from datetime import date
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill

from .checks import Issue
from .deadlines import deadline_diff, is_distributor_campaign
from .loader import PlanData

FILL_CHANGED = PatternFill("solid", fgColor="FFF2CC")   # изменённый дедлайн
FILL_RISK = PatternFill("solid", fgColor="F8CBAD")      # дедлайн в прошлом
COL = {"brief": 11, "creative": 12, "approval": 13}


def write_corrected_plan(
    plan: PlanData,
    issues: list[Issue],
    today: date,
    out_path: str | Path,
) -> Path:
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy(plan.source_path, out_path)

    wb = load_workbook(out_path)
    ws = None
    for sheet in wb.worksheets:
        if "контент-план" in sheet.title.lower():
            ws = sheet
            break
    if ws is None:
        raise KeyError("Не найдена вкладка контент-плана")

    issues_by_campaign: dict[str, list[Issue]] = {}
    for i in issues:
        for cid in i.campaign_ids:
            issues_by_campaign.setdefault(cid, []).append(i)

    for c in plan.campaigns:
        rule = plan.rule_for(c)
        offsets = {"brief": rule.brief, "creative": rule.creative, "approval": rule.approval}
        if is_distributor_campaign(c, plan) and c.campaign_type.strip() != "Retail support":
            offsets["approval"] += 1

        changed = {d["stage"] for d in deadline_diff(c)}
        for stage, col in COL.items():
            cell = ws.cell(row=c.row, column=col)
            cell.value = f"=WORKDAY(J{c.row}-{offsets[stage]}+1,-1)"
            cell.number_format = "DD.MM.YYYY"
            if stage in changed:
                cell.fill = FILL_CHANGED
            if c.corrected[stage] < today and not c.is_done:
                cell.fill = FILL_RISK

        # статус-колонка: комментарий с найденными проблемами
        found = issues_by_campaign.get(c.id, [])
        if found:
            status_cell = ws.cell(row=c.row, column=18)
            text = "\n".join(f"[{i.priority}] {i.title}" for i in found)
            status_cell.comment = Comment(text, "AI Content Plan Agent")

    # легенда под таблицей
    last = max(c.row for c in plan.campaigns) + 3
    ws.cell(row=last, column=1, value="Пересчитано агентом").font = Font(bold=True)
    ws.cell(
        row=last + 1,
        column=1,
        value=(
            "Дедлайны считаются по правилам вкладки «2. Правила дедлайнов» для своего типа "
            "кампании. Формула WORKDAY(старт−N+1;−1) переносит дедлайн с выходного на "
            "предыдущий рабочий день. Меняете дату старта в колонке J — дедлайны "
            "пересчитываются автоматически."
        ),
    )
    ws.cell(row=last + 2, column=1, value="Жёлтая заливка — дедлайн изменён относительно исходного файла.")
    ws.cell(row=last + 3, column=1, value="Оранжевая заливка — дедлайн в прошлом при незакрытом статусе.")
    ws.cell(row=last + 4, column=1, value=f"Дата расчёта: {today.strftime('%d.%m.%Y')}")

    wb.save(out_path)
    return out_path
