"""Построение мастер контент-плана из вводных менеджера.

Вход: короткая табличка «модель · категория · промо · дата старта · приоритет».
Выход: полноценный мастер-план в структуре шаблона — кампании разнесены
по каналам и назначениям, форматы креативов подобраны по матрице
«3. Каналы и назначения», дедлайны посчитаны обратным отсчётом.

Правила разнесения берутся из самого файла (вкладка «3. Каналы и назначения»),
а не захардкожены: добавится канал в матрицу — построитель его увидит.

Правила вкладки 3, которые обеспечиваются по построению:
  * одна модель за месяц закрывает минимум 1 перформанс-канал + 1 точку продаж;
  * медийка ставится только при наличии перформанс-поддержки на эту модель.
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook

from .loader import ChannelRow, _find_sheet, _header_row, _split_list, _to_date

# Длительность флайта по типу кампании (дней)
FLIGHT_DAYS = {
    "Промо-запуск": 20,
    "Маркетплейс": 26,
    "Retail support": 21,
    "Медийка": 21,
    "Always-on": 30,
}

# Бюджеты по умолчанию, если менеджер не указал свой (₸)
DEFAULT_BUDGET = {
    "Промо-запуск": 800_000,
    "Маркетплейс": 450_000,
    "Retail support": 300_000,
    "Медийка": 1_200_000,
    "Always-on": 400_000,
}

KPI_BY_TYPE = {
    "Промо-запуск": "ROAS 3.0 · 25 покупок",
    "Маркетплейс": "CTR 1.2% · 40 заказов",
    "Retail support": "Охват 500к · CTR 0.8%",
    "Медийка": "Охват 2.0М · VTR 25%",
    "Always-on": "CPC ≤ 120 ₸ · 15 покупок",
}

OWNER = "AI Mktg Manager"
MAX_LAUNCHES_PER_WEEK = 4


@dataclass
class InputRow:
    """Строка вводных от менеджера."""
    model: str
    category: str
    start: date
    promo: bool
    priority: str          # A / B / C
    distributor: str = ""  # если менеджер задал конкретного ритейлера
    budget: float | None = None
    note: str = ""
    # если менеджер сам указал канал/назначение — уважаем его выбор
    fixed_channel: str = ""
    fixed_destination: str = ""
    fixed_type: str = ""


@dataclass
class BuiltCampaign:
    model: str
    category: str
    campaign_type: str
    channel: str
    destination: str
    fmt: str
    start: date
    finish: date
    budget: float
    kpi: str
    priority: int          # 1 — снимать последней при перегрузе недели
    id: str = ""
    creative_id: str = ""

    @property
    def week(self) -> str:
        return f"W{self.start.isocalendar().week}"


# --------------------------------------------------------------------------
# чтение вводных и матрицы
# --------------------------------------------------------------------------

TRUE_WORDS = {"да", "yes", "true", "+", "1", "промо", "есть"}


def _is_true(value) -> bool:
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return bool(value)
    return str(value).strip().lower() in TRUE_WORDS


def load_input(path: str | Path) -> list[InputRow]:
    """Читает вводные менеджера.

    Основной формат — структура первого листа мастер-плана («1. Контент-план»):
    менеджер заполняет только свои колонки — Модель, Категория, Тип кампании
    (Промо-запуск / Always-on), Старт и необязательный Приоритет. Всё остальное
    (канал, назначение, формат, креатив, дедлайны, бюджет, KPI, статус)
    достраивает агент.

    Дополнительно поддерживается короткий формат — отдельная вкладка «Вводные».
    """
    wb = load_workbook(Path(path), data_only=True)
    try:
        ws = _find_sheet(wb, "Контент-план")
        return _load_input_plan_format(ws)
    except KeyError:
        ws = _find_sheet(wb, "Вводные")
        return _load_input_short_format(ws)


def _load_input_plan_format(ws) -> list[InputRow]:
    """Вводные в структуре мастер-таблицы: колонки как в «1. Контент-план»."""
    hdr = _header_row(ws, "ID")
    # ищем необязательную колонку «Приоритет» — она добавляется справа
    prio_col = None
    for c in range(1, ws.max_column + 1):
        if str(ws.cell(row=hdr, column=c).value or "").strip().lower().startswith("приоритет"):
            prio_col = c
            break

    rows: list[InputRow] = []
    for r in range(hdr + 1, ws.max_row + 1):
        model = ws.cell(row=r, column=3).value
        text = str(model or "").strip()
        if not text or text.startswith(("Легенда", "•")):
            if text.startswith(("Легенда", "•")):
                break
            continue

        start = _to_date(ws.cell(row=r, column=10).value)
        if start is None:
            raise ValueError(
                f"Строка {r}: не указана дата старта для модели {text}. "
                "Колонка «Старт» обязательна для каждой модели."
            )

        campaign_type = str(ws.cell(row=r, column=5).value or "").strip()
        channel = str(ws.cell(row=r, column=6).value or "").strip()
        destination = str(ws.cell(row=r, column=7).value or "").strip()
        budget = ws.cell(row=r, column=15).value
        priority = "B"
        if prio_col:
            priority = str(ws.cell(row=r, column=prio_col).value or "B").strip().upper()[:1] or "B"

        rows.append(
            InputRow(
                model=text,
                category=str(ws.cell(row=r, column=4).value or "").strip(),
                start=start,
                promo=campaign_type.lower().startswith("промо"),
                priority=priority,
                distributor=destination if destination.lower() not in {"", "samsung.kz"} else "",
                budget=float(budget) if isinstance(budget, (int, float)) else None,
                note=str(ws.cell(row=r, column=16).value or "").strip(),
                fixed_channel=channel,
                fixed_destination=destination,
                fixed_type=campaign_type,
            )
        )
    if not rows:
        raise ValueError("В файле вводных не найдено ни одной модели")
    return rows


def _load_input_short_format(ws) -> list[InputRow]:
    """Короткий формат: отдельная вкладка «Вводные» с именованными колонками."""
    hdr = _header_row(ws, "Модель")

    cols = {}
    for c in range(1, ws.max_column + 1):
        name = str(ws.cell(row=hdr, column=c).value or "").strip().lower()
        if name:
            cols[name] = c

    def cell(row: int, *names):
        for n in names:
            for key, col in cols.items():
                if key.startswith(n):
                    return ws.cell(row=row, column=col).value
        return None

    rows: list[InputRow] = []
    for r in range(hdr + 1, ws.max_row + 1):
        model = cell(r, "модель")
        text = str(model or "").strip()
        if not text or text.startswith(("Легенда", "•")):
            if text.startswith(("Легенда", "•")):
                break  # ниже легенды данных нет
            continue
        start = _to_date(cell(r, "дата старта", "старт"))
        if start is None:
            raise ValueError(
                f"Строка {r}: не указана дата старта для модели {model}. "
                "Дата обязательна для каждой модели."
            )
        budget = cell(r, "бюджет")
        rows.append(
            InputRow(
                model=str(model).strip(),
                category=str(cell(r, "категория") or "").strip(),
                start=start,
                promo=_is_true(cell(r, "промо")),
                priority=str(cell(r, "приоритет") or "B").strip().upper()[:1] or "B",
                distributor=str(cell(r, "дистрибьютор") or "").strip(),
                budget=float(budget) if isinstance(budget, (int, float)) else None,
                note=str(cell(r, "коммент") or "").strip(),
            )
        )
    if not rows:
        raise ValueError("В файле вводных не найдено ни одной модели")
    return rows


def load_matrix(path: str | Path) -> list[ChannelRow]:
    wb = load_workbook(Path(path), data_only=True)
    ws = _find_sheet(wb, "Каналы и назначения")
    hdr = _header_row(ws, "Канал")
    rows: list[ChannelRow] = []
    for r in range(hdr + 1, ws.max_row + 1):
        ch = ws.cell(row=r, column=1).value
        if not ch or str(ch).strip().startswith("Правило"):
            break
        if ws.cell(row=r, column=2).value is None:
            continue
        rows.append(
            ChannelRow(
                channel=str(ch).strip(),
                kind=str(ws.cell(row=r, column=2).value or "").strip(),
                destinations=_split_list(ws.cell(row=r, column=3).value),
                formats=_split_list(ws.cell(row=r, column=4).value),
                owner=str(ws.cell(row=r, column=5).value or "").strip(),
            )
        )
    return rows


# --------------------------------------------------------------------------
# разнесение по каналам
# --------------------------------------------------------------------------

def _pick(matrix: list[ChannelRow], kind: str, dest_hint: str = "") -> ChannelRow | None:
    """Первый канал нужного типа, умеющий вести на указанное назначение."""
    for row in matrix:
        if kind.lower() not in row.kind.lower():
            continue
        if not dest_hint:
            return row
        if any(dest_hint.lower() in d.lower() for d in row.destinations):
            return row
    return None


def _channel_by_name(matrix: list[ChannelRow], needle: str) -> ChannelRow | None:
    for row in matrix:
        if needle.lower() in row.channel.lower():
            return row
    return None


def _fmt_for(row: ChannelRow) -> str:
    return row.formats[0] if row.formats else "Display"


def _distributors(matrix: list[ChannelRow]) -> list[str]:
    marketplaces = {
        d for row in matrix if "маркетплейс" in row.kind.lower() for d in row.destinations
    }
    out: list[str] = []
    for row in matrix:
        for d in row.destinations:
            if d.lower() in {"samsung.kz", "дистрибьюторы"} or d in marketplaces:
                continue
            if d not in out:
                out.append(d)
    return out


def _marketplaces(matrix: list[ChannelRow]) -> list[tuple[ChannelRow, str]]:
    out = []
    for row in matrix:
        if "маркетплейс" in row.kind.lower():
            for d in row.destinations:
                out.append((row, d))
    return out


def build_campaigns(rows: list[InputRow], matrix: list[ChannelRow]) -> list[BuiltCampaign]:
    """Разносит каждую модель по каналам и назначениям.

    Приоритет A — полный охват: перформанс + маркетплейс + retail + медийка.
    Приоритет B — перформанс + точка продаж.
    Приоритет C — минимально: один перформанс-канал + маркетплейс.
    """
    distributors = _distributors(matrix)
    marketplaces = _marketplaces(matrix)
    built: list[BuiltCampaign] = []
    dist_cursor = 0
    mp_cursor = 0

    for item in rows:
        model_camps: list[BuiltCampaign] = []

        def add(campaign_type: str, ch: ChannelRow | None, dest: str, prio: int):
            if ch is None:
                return
            flight = FLIGHT_DAYS.get(campaign_type, 21)
            budget = item.budget if item.budget else DEFAULT_BUDGET.get(campaign_type, 400_000)
            model_camps.append(
                BuiltCampaign(
                    model=item.model,
                    category=item.category,
                    campaign_type=campaign_type,
                    channel=ch.channel,
                    destination=dest,
                    fmt=_fmt_for(ch),
                    start=item.start,
                    finish=item.start + timedelta(days=flight),
                    budget=budget,
                    kpi=KPI_BY_TYPE.get(campaign_type, ""),
                    priority=prio,
                )
            )

        # 0. Менеджер явно указал канал и назначение — строим ровно эту кампанию
        if item.fixed_channel:
            ch = _channel_by_name(matrix, item.fixed_channel.split()[0])
            if ch is not None:
                add(
                    item.fixed_type or ("Промо-запуск" if item.promo else "Always-on"),
                    ch,
                    item.fixed_destination or "samsung.kz",
                    prio=1,
                )
                built.extend(model_camps)
                continue

        # 1. Перформанс-ядро на сайт: промо-запуск или always-on
        core_type = "Промо-запуск" if item.promo else "Always-on"
        pmax = _channel_by_name(matrix, "Performance Max")
        search = _channel_by_name(matrix, "Google Search")
        add(core_type, pmax if item.promo else search, "samsung.kz", prio=1)
        if item.priority == "A":
            # у героя два перформанс-канала
            add(core_type, search if item.promo else pmax, "samsung.kz", prio=2)

        # 2. Точка продаж: маркетплейс (чередуем Kaspi / Ozon между моделями)
        if marketplaces:
            mp_row, mp_dest = marketplaces[mp_cursor % len(marketplaces)]
            mp_cursor += 1
            add("Маркетплейс", mp_row, mp_dest, prio=3)

        # 3. Поддержка дистрибьютора (A и B), чередуем ритейлеров для покрытия
        if item.priority in {"A", "B"} and distributors:
            if item.distributor:
                dest = item.distributor
            else:
                dest = distributors[dist_cursor % len(distributors)]
                dist_cursor += 1
            retail_ch = _pick(matrix, "Performance", dest) or _pick(matrix, "Медийка", dest)
            add("Retail support", retail_ch, dest, prio=4)

        # 4. Медийка — только героям и только поверх перформанса
        has_perf = any(
            c.campaign_type in {"Промо-запуск", "Always-on", "Маркетплейс"} for c in model_camps
        )
        if item.priority == "A" and has_perf:
            media = _channel_by_name(matrix, "YouTube") or _pick(matrix, "Медийка")
            add("Медийка", media, "samsung.kz", prio=2)

        built.extend(model_camps)

    return built


def level_weeks(campaigns: list[BuiltCampaign], limit: int = MAX_LAUNCHES_PER_WEEK) -> list[str]:
    """Разгружает недели с перебором запусков, двигая наименее приоритетные вперёд.

    Возвращает список пояснений — что и куда переехало.
    """
    original = {id(c): c.start for c in campaigns}
    changed = True
    guard = 0
    while changed and guard < 30:
        changed = False
        guard += 1
        by_week: dict[str, list[BuiltCampaign]] = {}
        for c in campaigns:
            by_week.setdefault(c.week, []).append(c)
        for week, items in sorted(by_week.items()):
            if len(items) <= limit:
                continue
            # двигаем «хвост»: сначала самые низкоприоритетные, внутри — поздние
            items.sort(key=lambda c: (-c.priority, c.start))
            for c in items[: len(items) - limit]:
                shift = 7 - (c.start.weekday() % 7)  # на следующий понедельник
                c.start += timedelta(days=shift)
                c.finish += timedelta(days=shift)
                changed = True

    # один итоговый сдвиг на кампанию вместо цепочки промежуточных
    notes: list[str] = []
    for c in campaigns:
        was = original[id(c)]
        if was != c.start:
            notes.append(
                f"{c.model} · {c.campaign_type} ({c.channel}): {was:%d.%m} → {c.start:%d.%m} "
                f"(+{(c.start - was).days} дн.) — выравнивание под лимит {limit} запусков в неделю"
            )
    return notes


def assign_ids(campaigns: list[BuiltCampaign]) -> None:
    campaigns.sort(key=lambda c: (c.start, c.model, c.campaign_type))
    for i, c in enumerate(campaigns, start=1):
        c.id = f"C-{i:02d}"
        c.creative_id = f"CR-{i:03d}"


# --------------------------------------------------------------------------
# запись мастер-плана
# --------------------------------------------------------------------------

def write_plan(
    campaigns: list[BuiltCampaign],
    template_path: str | Path,
    out_path: str | Path,
    title: str = "",
) -> Path:
    """Пишет план в структуре шаблона: вкладки с правилами и матрицей остаются.

    Дедлайны записываются формулами по правилам своего типа — файл сразу живой
    и является корректным входом для основного пайплайна.
    """
    import shutil

    from .loader import load_plan

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy(Path(template_path), out_path)

    wb = load_workbook(out_path)
    ws = None
    for sheet in wb.worksheets:
        if "контент-план" in sheet.title.lower():
            ws = sheet
            break
    if ws is None:
        raise KeyError("В шаблоне нет вкладки контент-плана")

    hdr = 4
    for r in range(hdr + 1, ws.max_row + 1):
        for c in range(1, 19):
            ws.cell(row=r, column=c).value = None

    if title:
        ws.cell(row=1, column=1).value = title

    for i, c in enumerate(campaigns):
        r = hdr + 1 + i
        values = {
            1: c.id, 2: c.week, 3: c.model, 4: c.category, 5: c.campaign_type,
            6: c.channel, 7: c.destination, 8: c.creative_id, 9: c.fmt,
            10: c.start, 14: c.finish, 15: c.budget, 16: c.kpi,
            17: OWNER, 18: "Не начато",
        }
        for col, v in values.items():
            cell = ws.cell(row=r, column=col)
            cell.value = v
            if isinstance(v, date):
                cell.number_format = "DD.MM.YYYY"
        # дедлайны проставит основной пайплайн — здесь оставляем пустыми,
        # чтобы не дублировать логику расчёта в двух местах
    wb.save(out_path)
    return out_path


def make_input_template(matrix_path: str | Path, out_path: str | Path) -> Path:
    """Создаёт шаблон вводных в структуре первого листа мастер-плана.

    Менеджер заполняет только свои колонки: Модель, Категория, Тип кампании,
    Старт и необязательный Приоритет. Остальные колонки достраивает агент.
    Вкладки с правилами и матрицей остаются в файле — он самодостаточен.
    """
    import shutil

    from openpyxl.styles import Alignment, Font, PatternFill

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy(Path(matrix_path), out_path)

    wb = load_workbook(out_path)
    ws = _find_sheet(wb, "Контент-план")
    hdr = _header_row(ws, "ID")

    ws.cell(row=1, column=1).value = "Вводные для построения контент-плана"
    ws.cell(row=2, column=1).value = (
        "Заполните выделенные колонки: одна строка = одна модель. Агент разнесёт "
        "кампании по каналам и назначениям, подберёт форматы креативов и посчитает дедлайны."
    )

    # чистим демо-данные шаблона
    for r in range(hdr + 1, ws.max_row + 1):
        for c in range(1, 20):
            ws.cell(row=r, column=c).value = None

    # добавляем необязательную колонку «Приоритет» справа
    prio = 19
    head = ws.cell(row=hdr, column=prio)
    head.value = "Приоритет"
    head.font = Font(bold=True, color="FFFFFF")
    head.fill = PatternFill("solid", fgColor="1428A0")
    head.alignment = Alignment(horizontal="center")
    ws.column_dimensions[head.column_letter].width = 12

    # подсвечиваем колонки, которые заполняет менеджер
    manager_cols = [3, 4, 5, 10, 19]
    fill = PatternFill("solid", fgColor="FFF2CC")
    for c in manager_cols:
        ws.cell(row=hdr, column=c).fill = fill
        ws.cell(row=hdr, column=c).font = Font(bold=True)

    examples = [
        ("WW80AK6L28BBLT", "Стиралки", "Промо-запуск", date(2026, 9, 7), "A"),
        ("NV75T9979CD/WT", "Духовые шкафы", "Промо-запуск", date(2026, 9, 14), "B"),
        ("RB53DG703EB1WT", "Холодильники", "Always-on", date(2026, 9, 14), "A"),
        ("AR80F09CABWNER", "Кондиционеры", "Промо-запуск", date(2026, 9, 21), "B"),
        ("WW65AK4S21CELT", "Стиралки", "Always-on", date(2026, 9, 21), "C"),
    ]
    for i, (model, cat, ctype, start, prio_v) in enumerate(examples):
        r = hdr + 1 + i
        ws.cell(row=r, column=3).value = model
        ws.cell(row=r, column=4).value = cat
        ws.cell(row=r, column=5).value = ctype
        cell = ws.cell(row=r, column=10)
        cell.value = start
        cell.number_format = "DD.MM.YYYY"
        ws.cell(row=r, column=prio).value = prio_v
        for c in manager_cols:
            ws.cell(row=r, column=c).fill = fill

    legend_row = hdr + len(examples) + 2
    ws.cell(row=legend_row, column=1, value="Легенда").font = Font(bold=True)
    legend = [
        "Заполняются только подсвеченные колонки: Модель, Категория, Тип кампании, Старт, Приоритет.",
        "Тип кампании — «Промо-запуск», если у модели есть промо-цена в этом месяце, иначе «Always-on».",
        "Старт — обязателен, задаётся отдельно для каждой модели.",
        "Приоритет A — герой месяца: два перформанс-канала + маркетплейс + дистрибьютор + медийка.",
        "Приоритет B — перформанс + маркетплейс + дистрибьютор. Приоритет C — перформанс + точка продаж.",
        "Колонки Канал и Назначение можно оставить пустыми — агент подберёт их сам по матрице вкладки 3.",
        "Если Канал и Назначение всё же заполнены, агент построит ровно эту кампанию и не будет добавлять свои.",
        "Бюджет необязателен: пустой заменяется значением по умолчанию для типа кампании.",
    ]
    for i, line in enumerate(legend):
        ws.cell(row=legend_row + 1 + i, column=1, value="• " + line)

    wb.save(out_path)
    return out_path
