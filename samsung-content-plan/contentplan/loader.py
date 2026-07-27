"""Загрузка мастер контент-плана и справочников из Excel.

Ключевой принцип: правила дедлайнов НЕ захардкожены в коде — они читаются
с вкладки «2. Правила дедлайнов». Если в следующем месяце маркетинг поменяет
сроки продакшена, пайплайн подхватит их без правки кода.
"""
from __future__ import annotations

import re
import warnings
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# Вкладки исходного файла (ищем по началу названия — устойчиво к «1. »/«1.»)
SHEET_PLAN = "Контент-план"
SHEET_RULES = "Правила дедлайнов"
SHEET_MATRIX = "Каналы и назначения"
SHEET_TRACKING = "Трекинг"

STATUS_DONE = "Готово"

# В файле используется типографский минус U+2212, а не дефис
_NUM_RE = re.compile(r"[−–—-]?\s*(\d+)")


def _to_days(cell) -> int:
    """'−14 дней' -> 14 (модуль сдвига назад от даты старта)."""
    if cell is None:
        return 0
    if isinstance(cell, (int, float)):
        return abs(int(cell))
    m = _NUM_RE.search(str(cell))
    return int(m.group(1)) if m else 0


def _to_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


@dataclass
class DeadlineRule:
    campaign_type: str
    brief: int          # дней до старта
    creative: int
    approval: int
    upload: int
    comment: str = ""

    @property
    def lead_time(self) -> int:
        """Сколько дней минимально нужно от брифа до старта."""
        return max(self.brief, self.creative, self.approval, self.upload)


@dataclass
class Campaign:
    row: int
    id: str
    week: str
    model: str
    category: str
    campaign_type: str
    channel: str
    destination: str
    creative_id: str
    fmt: str
    start: date
    finish: date | None
    budget: float
    kpi: str
    owner: str
    status: str
    # дедлайны из исходного файла (посчитанные единым правилом)
    original: dict[str, date | None] = field(default_factory=dict)
    # корректные дедлайны, посчитанные по правилам своего типа
    corrected: dict[str, date | None] = field(default_factory=dict)

    @property
    def is_done(self) -> bool:
        return str(self.status).strip().lower() == STATUS_DONE.lower()


@dataclass
class ChannelRow:
    channel: str
    kind: str                 # Performance / Медийка / Маркетплейс
    destinations: list[str]
    formats: list[str]
    owner: str


@dataclass
class Thresholds:
    """Пороги тревоги — читаются с вкладки «4. Трекинг» и «2. Правила»."""
    max_launches_per_week: int = 4
    max_deadlines_per_week: int = 6
    creative_ready_days: int = 7
    media_budget_share: float = 0.40


@dataclass
class PlanData:
    campaigns: list[Campaign]
    rules: dict[str, DeadlineRule]
    matrix: list[ChannelRow]
    thresholds: Thresholds
    title: str
    source_path: Path

    def rule_for(self, campaign: Campaign) -> DeadlineRule:
        rule = self.rules.get(campaign.campaign_type.strip())
        if rule is None:
            # неизвестный тип — берём самый консервативный набор правил
            rule = max(self.rules.values(), key=lambda r: r.lead_time)
        return rule

    @property
    def marketplaces(self) -> set[str]:
        out: set[str] = set()
        for row in self.matrix:
            if "маркетплейс" in row.kind.lower():
                out.update(row.destinations)
        return out

    @property
    def distributors(self) -> set[str]:
        """Дистрибьюторы = назначения из матрицы, кроме сайта и маркетплейсов."""
        own = {"samsung.kz", "дистрибьюторы"}
        out: set[str] = set()
        for row in self.matrix:
            for dest in row.destinations:
                if dest.lower() not in own and dest not in self.marketplaces:
                    out.add(dest)
        return out

    def channel_kind(self, channel: str) -> str:
        ch = channel.strip().lower()
        for row in self.matrix:
            if row.channel.strip().lower() == ch or ch.startswith(row.channel.split()[0].lower()):
                return row.kind
        return ""


def _find_sheet(wb, needle: str):
    for ws in wb.worksheets:
        if needle.lower() in ws.title.lower():
            return ws
    raise KeyError(f"Не найдена вкладка, содержащая «{needle}»")


def _header_row(ws, expected: str) -> int:
    for r in range(1, 15):
        if str(ws.cell(row=r, column=1).value or "").strip() == expected:
            return r
    return 4


def load_plan(path: str | Path, models_path: str | Path | None = None) -> PlanData:
    path = Path(path)
    wb_v = load_workbook(path, data_only=True)   # значения
    wb_f = load_workbook(path, data_only=False)  # формулы (для diff по исходным дедлайнам)

    ws = _find_sheet(wb_v, SHEET_PLAN)
    ws_f = _find_sheet(wb_f, SHEET_PLAN)
    hdr = _header_row(ws, "ID")

    campaigns: list[Campaign] = []
    for r in range(hdr + 1, ws.max_row + 1):
        cid = ws.cell(row=r, column=1).value
        if not cid or not str(cid).strip().upper().startswith("C-"):
            continue
        start = _to_date(ws.cell(row=r, column=10).value)
        if start is None:
            continue
        campaigns.append(
            Campaign(
                row=r,
                id=str(cid).strip(),
                week=str(ws.cell(row=r, column=2).value or "").strip(),
                model=str(ws.cell(row=r, column=3).value or "").strip(),
                category=str(ws.cell(row=r, column=4).value or "").strip(),
                campaign_type=str(ws.cell(row=r, column=5).value or "").strip(),
                channel=str(ws.cell(row=r, column=6).value or "").strip(),
                destination=str(ws.cell(row=r, column=7).value or "").strip(),
                creative_id=str(ws.cell(row=r, column=8).value or "").strip(),
                fmt=str(ws.cell(row=r, column=9).value or "").strip(),
                start=start,
                finish=_to_date(ws.cell(row=r, column=14).value),
                budget=float(ws.cell(row=r, column=15).value or 0),
                kpi=str(ws.cell(row=r, column=16).value or "").strip(),
                owner=str(ws.cell(row=r, column=17).value or "").strip(),
                status=str(ws.cell(row=r, column=18).value or "").strip(),
                original=_original_deadlines(ws, ws_f, r, start),
            )
        )

    rules = _load_rules(wb_v)
    matrix = _load_matrix(wb_v)
    thresholds = _load_thresholds(wb_v)

    return PlanData(
        campaigns=campaigns,
        rules=rules,
        matrix=matrix,
        thresholds=thresholds,
        title=str(ws.cell(row=1, column=1).value or "Контент-план"),
        source_path=path,
    )


def _original_deadlines(ws, ws_f, row: int, start: date) -> dict[str, date | None]:
    """Дедлайны, как они лежат в исходном файле.

    Формулы вида '=J5-14' openpyxl отдаёт без кэша значений, если файл писала
    не Excel — поэтому парсим смещение прямо из формулы, а если там значение,
    берём его как есть.
    """
    out: dict[str, date | None] = {}
    for key, col in (("brief", 11), ("creative", 12), ("approval", 13)):
        val = _to_date(ws.cell(row=row, column=col).value)
        if val is None:
            formula = ws_f.cell(row=row, column=col).value
            if isinstance(formula, str) and formula.startswith("="):
                m = re.search(r"-\s*(\d+)", formula)
                if m:
                    from datetime import timedelta
                    val = start - timedelta(days=int(m.group(1)))
        out[key] = val
    return out


def _load_rules(wb) -> dict[str, DeadlineRule]:
    ws = _find_sheet(wb, SHEET_RULES)
    hdr = _header_row(ws, "Тип кампании")
    rules: dict[str, DeadlineRule] = {}
    for r in range(hdr + 1, ws.max_row + 1):
        name = ws.cell(row=r, column=1).value
        if not name or str(name).strip().startswith("Дополнительные"):
            break
        if ws.cell(row=r, column=2).value is None:
            continue
        rules[str(name).strip()] = DeadlineRule(
            campaign_type=str(name).strip(),
            brief=_to_days(ws.cell(row=r, column=2).value),
            creative=_to_days(ws.cell(row=r, column=3).value),
            approval=_to_days(ws.cell(row=r, column=4).value),
            upload=_to_days(ws.cell(row=r, column=5).value),
            comment=str(ws.cell(row=r, column=6).value or "").strip(),
        )
    return rules


def _split_list(value) -> list[str]:
    if not value:
        return []
    return [p.strip() for p in re.split(r"[·•|]", str(value)) if p.strip()]


def _load_matrix(wb) -> list[ChannelRow]:
    ws = _find_sheet(wb, SHEET_MATRIX)
    hdr = _header_row(ws, "Канал")
    rows: list[ChannelRow] = []
    for r in range(hdr + 1, ws.max_row + 1):
        ch = ws.cell(row=r, column=1).value
        if not ch or str(ch).strip().startswith("Правило"):
            break
        if ws.cell(row=r, column=2).value is None:
            continue
        rows.append(
            ChannelRow(
                channel=str(ch).strip(),
                kind=str(ws.cell(row=r, column=2).value or "").strip(),
                destinations=_split_list(ws.cell(row=r, column=3).value),
                formats=_split_list(ws.cell(row=r, column=4).value),
                owner=str(ws.cell(row=r, column=5).value or "").strip(),
            )
        )
    return rows


def _load_thresholds(wb) -> Thresholds:
    """Вытаскиваем числовые пороги из текста вкладок «Трекинг» и «Правила»."""
    th = Thresholds()
    texts: list[str] = []
    for needle in (SHEET_TRACKING, SHEET_RULES):
        ws = _find_sheet(wb, needle)
        for row in ws.iter_rows(values_only=True):
            texts.extend(str(c) for c in row if isinstance(c, str))

    for t in texts:
        low = t.lower()
        m = re.search(r"больше\s+(\d+)", low)
        if m and "перегруз" in low:
            th.max_launches_per_week = int(m.group(1))
        elif m and ("не успеть" in low or "риск" in low):
            th.max_deadlines_per_week = int(m.group(1))
        m2 = re.search(r"не больше\s+(\d+)\s+запуск", low)
        if m2:
            th.max_launches_per_week = int(m2.group(1))
        m3 = re.search(r"медийка\s*>\s*(\d+)\s*%", low)
        if m3:
            th.media_budget_share = int(m3.group(1)) / 100
        m4 = re.search(r"старт[е]?\s*<\s*(\d+)\s*дн", low)
        if m4:
            th.creative_ready_days = int(m4.group(1))
    return th


def load_priority_models(path: str | Path) -> list[str]:
    """Список приоритетных моделей из файла Задания №1 (вкладка «Вводные»).

    Связка двух заданий: модели, под которые готовятся креативы, обязаны
    иметь хотя бы одну кампанию в контент-плане.
    """
    wb = load_workbook(Path(path), data_only=True)
    ws = _find_sheet(wb, "Вводные")
    hdr = _header_row(ws, "Tier")
    col = None
    for c in range(1, ws.max_column + 1):
        if str(ws.cell(row=hdr, column=c).value or "").strip().lower() == "model code":
            col = c
            break
    if col is None:
        return []
    models = []
    for r in range(hdr + 1, ws.max_row + 1):
        val = ws.cell(row=r, column=col).value
        if val and not str(val).strip().startswith("Легенда"):
            models.append(str(val).strip())
    return models
